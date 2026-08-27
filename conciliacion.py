# conciliacion.py
"""
Conciliación de cobranzas de Mercado Pago vs. sucursales.

Reglas:
1) Se carga UN archivo de Mercado Pago y TODOS los archivos de sucursales.
2) Fase 1: primero se intentan conciliar SOLO las filas de sucursal cuyo
   "Valores" sea MercadoPago.
3) Fase 2: después se recorren TODAS las filas restantes de las sucursales,
   sin importar el medio de pago escrito en "Valores".
4) Una operación de Mercado Pago se puede usar UNA SOLA VEZ en todo el proceso,
   entre todas las sucursales.
5) La coincidencia es por FECHA + IMPORTE exactos (TOLERANCIA_DIAS = 0).
6) Si una fila de sucursal encuentra MP:
      Resultado = "Mercado Pago"
      Operación MP = operation_id
   Si no encuentra:
      Resultado = valor original de "Valores"
7) En Mercado Pago:
      Resultado = "ok" si la operación fue usada
      Resultado = "no encontrado" si quedó sin usar.
8) Los Excel originales NO se modifican. Se generan copias *_conciliado.xlsx.
"""

import os
import re
import sys
import unicodedata
from collections import defaultdict, deque
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN
# ============================================================

TOLERANCIA_DIAS = 0
SUFIJO_SALIDA = "_conciliado.xlsx"

CANDIDATOS_MP_ID = [
    "numero de operacion",
    "nro de operacion",
    "operation id",
    "operation_id",
    "operacion",
]
CANDIDATOS_MP_FECHA = [
    "fecha de compra",
    "date_created",
    "fecha",
]
CANDIDATOS_MP_IMPORTE = [
    "valor del producto",
    "transaction_amount",
    "importe",
    "monto",
]

CANDIDATOS_SUC_VALORES = ["valores", "medio de pago", "medio"]
CANDIDATOS_SUC_FECHA = ["fecha"]
CANDIDATOS_SUC_IMPORTE = ["importe", "monto"]


# ============================================================
# CONSOLA CON COLORES
# ============================================================

COLOR = sys.stdout.isatty()

class C:
    VERDE = "\033[92m" if COLOR else ""
    ROJO = "\033[91m" if COLOR else ""
    AMARILLO = "\033[93m" if COLOR else ""
    CYAN = "\033[96m" if COLOR else ""
    MAGENTA = "\033[95m" if COLOR else ""
    BOLD = "\033[1m" if COLOR else ""
    RESET = "\033[0m" if COLOR else ""


def log(msg=""):
    print(msg, flush=True)


def ok(msg):
    log(f"{C.VERDE}OK  {msg}{C.RESET}")


def bad(msg):
    log(f"{C.ROJO}NO  {msg}{C.RESET}")


def warn(msg):
    log(f"{C.AMARILLO}!!  {msg}{C.RESET}")


def info(msg):
    log(f"{C.CYAN}{msg}{C.RESET}")


def titulo(msg):
    log(
        f"\n{C.BOLD}{C.CYAN}"
        f"{'=' * 78}\n{msg}\n{'=' * 78}"
        f"{C.RESET}"
    )


# ============================================================
# NORMALIZACIÓN
# ============================================================

def normalizar_texto(valor) -> str:
    if valor is None:
        return ""
    s = str(valor).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[\s_.\-]+", "", s)
    return s


def encontrar_columna(df, candidatos):
    """
    Busca una columna por nombre normalizado.
    Primero intenta igualdad exacta y después coincidencia contenida.
    """
    columnas = list(df.columns)
    normalizadas = {c: normalizar_texto(c) for c in columnas}

    # Igualdad exacta
    for candidato in candidatos:
        cn = normalizar_texto(candidato)
        for col in columnas:
            if normalizadas[col] == cn:
                return col

    # Coincidencia contenida
    for candidato in candidatos:
        cn = normalizar_texto(candidato)
        coincidencias = [
            col for col in columnas if cn and cn in normalizadas[col]
        ]
        if coincidencias:
            return coincidencias[0]

    return None


def a_fecha(valor):
    if pd.isna(valor):
        return None

    if isinstance(valor, (datetime, date)):
        return pd.Timestamp(valor).normalize()

    try:
        f = pd.to_datetime(valor, dayfirst=True, errors="coerce")
        if pd.isna(f):
            return None
        return f.normalize()
    except Exception:
        return None


def a_importe(valor):
    if pd.isna(valor):
        return None

    if isinstance(valor, (int, float)):
        if pd.isna(valor):
            return None
        return round(float(valor), 2)

    s = str(valor).strip()
    if not s:
        return None

    # Elimina símbolos de moneda y espacios.
    s = re.sub(r"[^\d,.\-]", "", s)

    # 12.345,67 -> 12345.67
    # 12345,67 -> 12345.67
    # 12,345.67 -> 12345.67
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")

    try:
        return round(float(s), 2)
    except Exception:
        return None


def es_mercadopago(valor):
    return normalizar_texto(valor) in {
        "mercadopago",
        "mercadopagoqr",
        "mp",
    }


def clave(fecha, importe):
    if fecha is None or importe is None:
        return None
    return fecha, importe


# ============================================================
# POOL GLOBAL DE MERCADO PAGO
# ============================================================

class PoolMP:
    """
    Contiene TODAS las operaciones de MP de TODAS las hojas.

    La estructura indice permite que, para una clave (fecha, importe),
    existan N operaciones. Cada operación se saca de la cola una sola vez.

    Esto es lo que impide que, por ejemplo, si MP tiene 3 cobros de
    $10.000 el mismo día, se asignen 4 o 5 filas de sucursales.
    """

    def __init__(self, hojas_mp):
        self.hojas = hojas_mp
        self.indice = defaultdict(deque)
        self.total = 0
        self.usadas = 0

        for nombre_hoja, meta in self.hojas.items():
            df = meta["df"]
            fechas = df[meta["col_fecha"]].map(a_fecha)
            importes = df[meta["col_importe"]].map(a_importe)

            meta["resultado"] = [""] * len(df)
            meta["usado_en"] = [""] * len(df)

            for idx in range(len(df)):
                k = clave(fechas.iat[idx], importes.iat[idx])
                if k is not None:
                    self.indice[k].append((nombre_hoja, idx))

            self.total += len(df)

    def buscar(self, fecha, importe):
        """
        Devuelve UNA operación disponible y la consume.
        Si no hay ninguna disponible, devuelve None.
        """
        if fecha is None or importe is None:
            return None

        fechas_a_probar = []
        for delta in range(TOLERANCIA_DIAS + 1):
            if delta == 0:
                fechas_a_probar.append(fecha)
            else:
                fechas_a_probar.append(
                    fecha - pd.Timedelta(days=delta)
                )
                fechas_a_probar.append(
                    fecha + pd.Timedelta(days=delta)
                )

        for f in fechas_a_probar:
            dq = self.indice.get((f, importe))
            if not dq:
                continue

            while dq:
                nombre_hoja, idx = dq.popleft()

                meta = self.hojas[nombre_hoja]
                if meta["resultado"][idx] != "":
                    # Seguridad extra: nunca reutilizar una operación.
                    continue

                meta["resultado"][idx] = "ok"
                self.usadas += 1

                operation_id = meta["df"].at[idx, meta["col_id"]]
                return {
                    "hoja": nombre_hoja,
                    "idx": idx,
                    "operation_id": operation_id,
                }

        return None

    def cerrar_no_usadas(self):
        cantidad = 0
        for meta in self.hojas.values():
            for i in range(len(meta["resultado"])):
                if meta["resultado"][i] == "":
                    meta["resultado"][i] = "no encontrado"
                    cantidad += 1
        return cantidad


# ============================================================
# CARGA
# ============================================================

def cargar_mp(path, logger=info):
    logger(f"Cargando Mercado Pago: {path}")

    hojas_raw = pd.read_excel(path, sheet_name=None)
    hojas = {}

    for nombre, df in hojas_raw.items():
        if df.empty:
            warn(f"Hoja MP '{nombre}' vacía. Se omite.")
            continue

        col_id = encontrar_columna(df, CANDIDATOS_MP_ID)
        col_fecha = encontrar_columna(df, CANDIDATOS_MP_FECHA)
        col_importe = encontrar_columna(df, CANDIDATOS_MP_IMPORTE)

        faltantes = []
        if col_id is None:
            faltantes.append("numero de operacion")
        if col_fecha is None:
            faltantes.append("fecha")
        if col_importe is None:
            faltantes.append("importe")

        if faltantes:
            bad(
                f"Hoja MP '{nombre}': faltan columnas {faltantes}. "
                f"Se omite."
            )
            continue

        hojas[nombre] = {
            "df": df.reset_index(drop=True),
            "col_id": col_id,
            "col_fecha": col_fecha,
            "col_importe": col_importe,
        }

        info(
            f"  MP / {nombre}: {len(df):,} filas | "
            f"id='{col_id}' | fecha='{col_fecha}' | importe='{col_importe}'"
        )

    return hojas


def cargar_sucursal(path, logger=info):
    logger(f"Cargando sucursal: {path}")

    hojas_raw = pd.read_excel(path, sheet_name=None)
    hojas = {}

    for nombre, df in hojas_raw.items():
        if df.empty:
            warn(f"Hoja '{nombre}' vacía. Se omite.")
            continue

        col_valores = encontrar_columna(df, CANDIDATOS_SUC_VALORES)
        col_fecha = encontrar_columna(df, CANDIDATOS_SUC_FECHA)
        col_importe = encontrar_columna(df, CANDIDATOS_SUC_IMPORTE)

        faltantes = []
        if col_valores is None:
            faltantes.append("Valores")
        if col_fecha is None:
            faltantes.append("Fecha")
        if col_importe is None:
            faltantes.append("Importe")

        if faltantes:
            bad(
                f"Hoja '{nombre}' de '{Path(path).name}': "
                f"faltan columnas {faltantes}. Se omite."
            )
            continue

        df = df.reset_index(drop=True)

        hojas[nombre] = {
            "df": df,
            "col_valores": col_valores,
            "col_fecha": col_fecha,
            "col_importe": col_importe,
            "fecha_norm": df[col_fecha].map(a_fecha),
            "importe_norm": df[col_importe].map(a_importe),
            "resultado": [""] * len(df),
            "operacion_mp": [""] * len(df),
        }

        info(
            f"  {Path(path).name} / {nombre}: {len(df):,} filas | "
            f"Valores='{col_valores}' | fecha='{col_fecha}' | "
            f"importe='{col_importe}'"
        )

    return hojas


# ============================================================
# FASE 1
# ============================================================

def fase_1(sucursales, pool, logger=print):
    titulo("FASE 1 — PRIORIZAR filas ya marcadas como MercadoPago")

    total_candidatas = 0
    total_encontradas = 0

    for archivo, hojas in sucursales.items():
        for hoja, meta in hojas.items():
            df = meta["df"]
            # Vectorizado con pandas en vez de un loop fila-por-fila con
            # .at (que en archivos de 50.000+ filas es notablemente
            # más lento que un .map sobre toda la columna de una vez).
            es_mp_mask = df[meta["col_valores"]].map(es_mercadopago)
            indices = df.index[es_mp_mask].tolist()

            total_candidatas += len(indices)
            encontradas = 0

            for i in indices:
                fecha = meta["fecha_norm"].iat[i]
                importe = meta["importe_norm"].iat[i]
                match = pool.buscar(fecha, importe)

                if match is not None:
                    meta["resultado"][i] = "Mercado Pago"
                    meta["operacion_mp"][i] = str(match["operation_id"])
                    encontradas += 1
                    total_encontradas += 1

                    logger(
                        f"{C.VERDE}✓ F1 | {Path(archivo).name} / {hoja} "
                        f"| fila {i + 2} | MercadoPago | "
                        f"{fecha.date() if fecha is not None else '?'} | "
                        f"${importe:,.2f} | "
                        f"MP: {match['hoja']} fila {match['idx'] + 2} | "
                        f"op. {match['operation_id']} "
                        f"{C.RESET}"
                    )
                else:
                    # No se marca todavía: queda disponible para Fase 2.
                    logger(
                        f"{C.ROJO}✗ F1 | {Path(archivo).name} / {hoja} "
                        f"| fila {i + 2} | dice MercadoPago pero no hay "
                        f"operación MP disponible con esa fecha/importe "
                        f"(se reintenta en F2){C.RESET}"
                    )

            if indices:
                info(
                    f"{Path(archivo).name} / {hoja}: "
                    f"{encontradas:,}/{len(indices):,} confirmadas en F1."
                )

    info(
        f"TOTAL FASE 1: {total_encontradas:,} confirmadas de "
        f"{total_candidatas:,} filas que ya decían MercadoPago."
    )


# ============================================================
# FASE 2
# ============================================================

def fase_2(sucursales, pool, logger=print):
    titulo(
        "FASE 2 — BUSCAR MP ESCONDIDOS EN TODAS LAS FILAS RESTANTES"
    )

    total_revisadas = 0
    total_encontradas = 0

    for archivo, hojas in sucursales.items():
        for hoja, meta in hojas.items():
            df = meta["df"]

            pendientes = [
                i for i in range(len(df))
                if meta["resultado"][i] == ""
            ]

            encontradas = 0

            for i in pendientes:
                total_revisadas += 1

                fecha = meta["fecha_norm"].iat[i]
                importe = meta["importe_norm"].iat[i]
                valor_original = df.at[i, meta["col_valores"]]

                match = pool.buscar(fecha, importe)

                if match is not None:
                    meta["resultado"][i] = "Mercado Pago"
                    meta["operacion_mp"][i] = str(match["operation_id"])
                    encontradas += 1
                    total_encontradas += 1

                    logger(
                        f"{C.VERDE}✓ F2 | {Path(archivo).name} / {hoja} "
                        f"| fila {i + 2} | estaba '{valor_original}' → "
                        f"Mercado Pago | "
                        f"{fecha.date() if fecha is not None else '?'} | "
                        f"${importe:,.2f} | "
                        f"MP: {match['hoja']} fila {match['idx'] + 2} | "
                        f"op. {match['operation_id']} "
                        f"{C.RESET}"
                    )
                else:
                    meta["resultado"][i] = valor_original

            info(
                f"{Path(archivo).name} / {hoja}: "
                f"{encontradas:,} MP escondidos encontrados de "
                f"{len(pendientes):,} filas restantes."
            )

    info(
        f"TOTAL FASE 2: {total_encontradas:,} cobros que estaban "
        f"clasificados con otro medio y fueron identificados como MP."
    )


# ============================================================
# VALIDACIÓN DE DUPLICADOS
# ============================================================

def validar_limite_mp(sucursales, pool):
    """
    Verifica que para cada (fecha, importe), la cantidad de filas
    conciliadas en sucursales nunca sea mayor que la cantidad de
    operaciones MP disponibles para esa misma clave.

    En condiciones normales siempre debe dar 0 violaciones porque
    PoolMP consume cada operación una sola vez.
    """
    titulo("CONTROL DE DUPLICADOS")

    conciliadas = defaultdict(int)

    for archivo, hojas in sucursales.items():
        for hoja, meta in hojas.items():
            for i, resultado in enumerate(meta["resultado"]):
                if resultado != "Mercado Pago":
                    continue

                k = clave(
                    meta["fecha_norm"].iat[i],
                    meta["importe_norm"].iat[i]
                )
                if k is not None:
                    conciliadas[k] += 1

    disponibles = defaultdict(int)
    for nombre, meta in pool.hojas.items():
        df = meta["df"]
        fechas = df[meta["col_fecha"]].map(a_fecha)
        importes = df[meta["col_importe"]].map(a_importe)

        for i in range(len(df)):
            k = clave(fechas.iat[i], importes.iat[i])
            if k is not None:
                disponibles[k] += 1

    violaciones = []
    for k, cantidad_suc in conciliadas.items():
        cantidad_mp = disponibles.get(k, 0)
        if cantidad_suc > cantidad_mp:
            violaciones.append((k, cantidad_suc, cantidad_mp))

    if not violaciones:
        ok(
            "CONTROL OK: ninguna combinación fecha + importe fue "
            "conciliada más veces que las operaciones disponibles en MP."
        )
    else:
        bad(
            f"SE ENCONTRARON {len(violaciones)} VIOLACIONES DE DUPLICADO."
        )
        for k, cs, cm in violaciones[:20]:
            bad(
                f"  {k[0].date()} | ${k[1]:,.2f} | "
                f"sucursales={cs} | MP={cm}"
            )

    return violaciones


# ============================================================
# GUARDADO
# ============================================================

def destino_conciliado(path, carpeta_base):
    """
    Arma la ruta de salida dentro de una subcarpeta "Conciliados"
    ubicada en carpeta_base (la carpeta del proyecto), NO en la
    carpeta donde esté físicamente el archivo original. Esto evita
    que, por ejemplo, el archivo de Mercado Pago (bajado a Descargas)
    termine generando su propia carpeta "Conciliados" separada de la
    de las sucursales.
    """
    p = Path(path)

    carpeta_destino = Path(carpeta_base) / "Conciliados"
    carpeta_destino.mkdir(parents=True, exist_ok=True)

    if p.suffix.lower() != ".xlsx":
        return carpeta_destino / (p.name + SUFIJO_SALIDA)
    return carpeta_destino / (p.stem + SUFIJO_SALIDA)


def escribir_sin_perder_formato(path, hojas, carpeta_base):
    """
    Abre el Excel original con openpyxl y solamente agrega/reemplaza
    las columnas Resultado y Operación MP al final de cada hoja.
    """
    destino = destino_conciliado(path, carpeta_base)

    # Copia primero el archivo completo y después modifica la copia.
    import shutil
    shutil.copy2(path, destino)

    wb = load_workbook(destino)

    for nombre, meta in hojas.items():
        if nombre not in wb.sheetnames:
            continue

        ws = wb[nombre]

        # Siempre se agregan al FINAL real de la hoja.
        col_resultado = ws.max_column + 1
        col_operacion = ws.max_column + 2

        ws.cell(row=1, column=col_resultado, value="Resultado")
        ws.cell(row=1, column=col_operacion, value="Operación MP")

        for i in range(len(meta["df"])):
            fila = i + 2
            ws.cell(
                row=fila,
                column=col_resultado,
                value=meta["resultado"][i]
            )

            op = meta["operacion_mp"][i]
            ws.cell(
                row=fila,
                column=col_operacion,
                value=op if op else None
            )

    wb.save(destino)
    return str(destino)


def escribir_mp(path, pool, carpeta_base):
    destino = destino_conciliado(path, carpeta_base)

    import shutil
    shutil.copy2(path, destino)

    wb = load_workbook(destino)

    for nombre, meta in pool.hojas.items():
        if nombre not in wb.sheetnames:
            continue

        ws = wb[nombre]
        col_resultado = ws.max_column + 1
        ws.cell(row=1, column=col_resultado, value="Resultado")

        for i, resultado in enumerate(meta["resultado"]):
            ws.cell(row=i + 2, column=col_resultado, value=resultado)

    wb.save(destino)
    return str(destino)


# ============================================================
# PROCESO PRINCIPAL
# ============================================================

def ejecutar(archivo_mp, archivos_sucursales, logger=print):
    """
    Ejecuta todo el proceso.

    archivo_mp: ruta al Excel de Mercado Pago.
    archivos_sucursales: lista con las rutas de las sucursales.
    """
    if not archivo_mp:
        raise ValueError("No se seleccionó el archivo de Mercado Pago.")

    if not archivos_sucursales:
        raise ValueError("No se seleccionó ningún archivo de sucursal.")

    archivo_mp = str(Path(archivo_mp).resolve())
    archivos_sucursales = [
        str(Path(p).resolve()) for p in archivos_sucursales
    ]

    titulo("CONCILIACIÓN MERCADO PAGO vs. SUCURSALES")
    info(f"Archivo MP: {archivo_mp}")
    info(f"Sucursales seleccionadas: {len(archivos_sucursales)}")

    # --------------------------------------------------------
    # Carga MP
    # --------------------------------------------------------
    titulo("CARGA DE MERCADO PAGO")
    hojas_mp = cargar_mp(archivo_mp)

    if not hojas_mp:
        raise RuntimeError(
            "No se pudo cargar ninguna hoja válida del archivo de MP."
        )

    pool = PoolMP(hojas_mp)
    info(f"Operaciones MP cargadas: {pool.total:,}")

    # --------------------------------------------------------
    # Carga sucursales
    # --------------------------------------------------------
    titulo("CARGA DE SUCURSALES")
    sucursales = {}

    for path in archivos_sucursales:
        hojas = cargar_sucursal(path)
        if hojas:
            sucursales[path] = hojas

    if not sucursales:
        raise RuntimeError(
            "No se pudo cargar ninguna hoja válida de sucursal."
        )

    # --------------------------------------------------------
    # IMPORTANTE:
    # Fase 1 COMPLETA antes de Fase 2.
    # Así las filas que ya dicen MercadoPago tienen prioridad
    # sobre los posibles MP escondidos.
    # --------------------------------------------------------
    fase_1(sucursales, pool, logger=logger)
    fase_2(sucursales, pool, logger=logger)

    # --------------------------------------------------------
    # Cierre y controles
    # --------------------------------------------------------
    no_encontradas = pool.cerrar_no_usadas()

    # Mostrar algunos casos concretos que quedaron sin encontrar.
    # No se imprimen decenas de miles de líneas si el archivo es grande.
    MAX_DETALLE_NO_ENCONTRADO = 200
    mostradas = 0
    for nombre, meta in pool.hojas.items():
        df_mp = meta["df"]
        for i, resultado in enumerate(meta["resultado"]):
            if resultado != "no encontrado":
                continue
            if mostradas < MAX_DETALLE_NO_ENCONTRADO:
                fecha = a_fecha(df_mp.at[i, meta["col_fecha"]])
                importe = a_importe(df_mp.at[i, meta["col_importe"]])
                logger(
                    f"{C.ROJO}✗ MP NO ENCONTRADO | {nombre} / fila {i + 2} "
                    f"| op. {df_mp.at[i, meta['col_id']]} | "
                    f"{fecha.date() if fecha is not None else '?'} | "
                    f"${importe:,.2f} | no apareció en ninguna sucursal "
                    f"{C.RESET}"
                )
                mostradas += 1

    if no_encontradas > MAX_DETALLE_NO_ENCONTRADO:
        warn(
            f"Se muestran solo los primeros {MAX_DETALLE_NO_ENCONTRADO} "
            f"casos rojos de MP no encontrado; el total es "
            f"{no_encontradas:,}."
        )

    violaciones = validar_limite_mp(sucursales, pool)

    # --------------------------------------------------------
    # Resumen
    # --------------------------------------------------------
    titulo("RESUMEN")

    total_filas = 0
    total_mp_suc = 0

    for archivo, hojas in sucursales.items():
        for hoja, meta in hojas.items():
            n = len(meta["df"])
            nmp = sum(
                1 for r in meta["resultado"]
                if r == "Mercado Pago"
            )

            total_filas += n
            total_mp_suc += nmp

            info(
                f"{Path(archivo).name} / {hoja}: "
                f"{n:,} filas | {nmp:,} Mercado Pago"
            )

    info(f"Total filas sucursales: {total_filas:,}")
    info(f"Total conciliado como Mercado Pago: {total_mp_suc:,}")
    info(f"Operaciones MP usadas: {pool.usadas:,}/{pool.total:,}")

    if no_encontradas:
        bad(
            f"{no_encontradas:,} operaciones de MP quedaron "
            f"como 'no encontrado'."
        )
    else:
        ok("Todas las operaciones de MP fueron encontradas.")

    if violaciones:
        bad(
            "ATENCIÓN: el control de duplicados encontró violaciones."
        )
    else:
        ok("Control de duplicados: OK.")

    # --------------------------------------------------------
    # Guardar copias
    # --------------------------------------------------------
    titulo("GUARDANDO ARCHIVOS")

    # La carpeta "del proyecto" se toma de donde están las sucursales
    # (no del archivo de MP, que muchas veces está en Descargas).
    # Así, MP y todas las sucursales van a UNA sola carpeta "Conciliados".
    carpeta_base = Path(archivos_sucursales[0]).parent
    info(f"Carpeta de salida: {carpeta_base / 'Conciliados'}")

    salidas = []

    salida_mp = escribir_mp(archivo_mp, pool, carpeta_base)
    salidas.append(salida_mp)
    ok(f"MP: {salida_mp}")

    for archivo, hojas in sucursales.items():
        salida = escribir_sin_perder_formato(archivo, hojas, carpeta_base)
        salidas.append(salida)
        ok(f"Sucursal: {salida}")

    titulo("PROCESO TERMINADO")
    ok("Los archivos originales NO fueron modificados.")

    return {
        "salidas": salidas,
        "operaciones_mp": pool.total,
        "operaciones_mp_usadas": pool.usadas,
        "no_encontradas": no_encontradas,
        "filas_sucursales": total_filas,
        "mp_en_sucursales": total_mp_suc,
        "violaciones_duplicado": len(violaciones),
    }


# ============================================================
# INTERFAZ GRÁFICA
# ============================================================

def iniciar_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext

    root = tk.Tk()
    root.title("Conciliación Mercado Pago vs. Sucursales")
    root.geometry("1050x700")
    root.minsize(850, 600)

    archivo_mp_var = tk.StringVar(value="")
    sucursales = []

    # ---------- helpers ----------
    # LOTE DE LOGS: con archivos grandes (50.000+ filas), Fase 1/Fase 2
    # pueden generar decenas de miles de líneas. Actualizar el widget
    # de Tkinter Y forzar redibujado (update_idletasks) en CADA línea
    # es, por lejos, el cuello de botella más grande del programa.
    # Por eso acá se juntan las líneas en un buffer y se vuelcan al
    # widget de a lotes (cada LOG_BATCH_SIZE líneas), reduciendo miles
    # de "reconfigurar + redibujar" a un puñado.
    LOG_BATCH_SIZE = 500
    _log_buffer = []

    def flush_log():
        if not _log_buffer:
            return
        consola.configure(state="normal")
        for texto, tag in _log_buffer:
            consola.insert("end", texto + "\n", tag)
        _log_buffer.clear()
        consola.see("end")
        consola.configure(state="disabled")
        root.update_idletasks()

    def agregar_log(texto, tag=None):
        _log_buffer.append((texto, tag))
        if len(_log_buffer) >= LOG_BATCH_SIZE:
            flush_log()

    def seleccionar_mp():
        path = filedialog.askopenfilename(
            title="Seleccionar archivo de Mercado Pago",
            filetypes=[
                ("Excel", "*.xlsx"),
                ("Excel", "*.xls"),
                ("Todos", "*.*"),
            ],
        )
        if path:
            archivo_mp_var.set(path)

    def seleccionar_sucursales():
        # Permite seleccionar VARIOS archivos de una sola vez.
        # Importante: NO borra los que ya estaban agregados.
        paths = filedialog.askopenfilenames(
            title="Seleccionar archivos de las sucursales",
            filetypes=[
                ("Excel", "*.xlsx"),
                ("Excel", "*.xls"),
                ("Todos", "*.*"),
            ],
        )
        if not paths:
            return

        for path in paths:
            if path not in sucursales:
                sucursales.append(path)
        actualizar_lista()

    def agregar_sucursal():
        # Alternativa más simple: agregar una sucursal por vez.
        path = filedialog.askopenfilename(
            title="Agregar archivo de una sucursal",
            filetypes=[
                ("Excel", "*.xlsx"),
                ("Excel", "*.xls"),
                ("Todos", "*.*"),
            ],
        )
        if path and path not in sucursales:
            sucursales.append(path)
            actualizar_lista()

    def quitar_sucursal():
        seleccion = lista.curselection()
        if not seleccion:
            messagebox.showwarning(
                "Seleccioná una sucursal",
                "Seleccioná primero una sucursal de la lista para quitarla."
            )
            return

        # Quitar desde abajo para no alterar los índices restantes.
        for indice in reversed(seleccion):
            del sucursales[indice]
        actualizar_lista()

    def actualizar_lista():
        lista.delete(0, "end")
        for i, path in enumerate(sucursales, 1):
            lista.insert(
                "end",
                f"{i}. {Path(path).name}"
            )

        etiqueta_cantidad.config(
            text=f"{len(sucursales)} archivo(s) de sucursal seleccionado(s)"
        )

    def limpiar_sucursales():
        sucursales.clear()
        actualizar_lista()

    def limpiar_consola():
        consola.configure(state="normal")
        consola.delete("1.0", "end")
        consola.configure(state="disabled")

    def ejecutar_gui():
        if not archivo_mp_var.get():
            messagebox.showwarning(
                "Falta Mercado Pago",
                "Seleccioná el archivo de Mercado Pago."
            )
            return

        if not sucursales:
            messagebox.showwarning(
                "Faltan sucursales",
                "Seleccioná los archivos de las sucursales."
            )
            return

        btn_ejecutar.config(state="disabled")
        btn_mp.config(state="disabled")
        btn_suc.config(state="disabled")

        limpiar_consola()

        def logger(msg):
            # Las funciones de conciliación mandan líneas con ANSI.
            # En la GUI no usamos ANSI: mostramos texto plano y damos
            # color según el prefijo.
            texto = re.sub(r"\x1b\[[0-9;]*m", "", str(msg))

            tag = None
            if "✓" in texto or texto.startswith("OK"):
                tag = "verde"
            elif "✗" in texto or texto.startswith("NO"):
                tag = "rojo"
            elif texto.startswith("!!"):
                tag = "amarillo"
            elif "==" in texto:
                tag = "titulo"

            agregar_log(texto, tag)

        try:
            resultado = ejecutar(
                archivo_mp_var.get(),
                list(sucursales),
                logger=logger,
            )

            flush_log()

            messagebox.showinfo(
                "Conciliación terminada",
                "Proceso terminado correctamente.\n\n"
                f"MP encontradas: {resultado['operaciones_mp_usadas']:,}\n"
                f"MP no encontradas: {resultado['no_encontradas']:,}\n"
                f"MP en sucursales: {resultado['mp_en_sucursales']:,}\n\n"
                "Los archivos *_conciliado.xlsx fueron generados "
                "junto a los originales."
            )

        except Exception as exc:
            agregar_log("")
            agregar_log("ERROR: " + str(exc), "rojo")
            flush_log()
            messagebox.showerror(
                "Error durante la conciliación",
                str(exc)
            )

        finally:
            # Red de seguridad: garantiza que cualquier línea que haya
            # quedado en el buffer se vuelque igual, pase lo que pase.
            flush_log()
            btn_ejecutar.config(state="normal")
            btn_mp.config(state="normal")
            btn_suc.config(state="normal")

    # ---------- layout ----------
    marco = tk.Frame(root, padx=15, pady=15)
    marco.pack(fill="both", expand=True)

    tk.Label(
        marco,
        text="Conciliación Mercado Pago vs. Sucursales",
        font=("TkDefaultFont", 16, "bold"),
    ).pack(anchor="w")

    tk.Label(
        marco,
        text=(
            "Primero confirma los que ya dicen MercadoPago. "
            "Después busca MP escondidos en TODAS las filas. "
            "Cada operación MP puede utilizarse una sola vez."
        ),
        justify="left",
    ).pack(anchor="w", pady=(5, 15))

    # MP
    fila_mp = tk.Frame(marco)
    fila_mp.pack(fill="x", pady=4)

    btn_mp = tk.Button(
        fila_mp,
        text="Seleccionar Mercado Pago",
        command=seleccionar_mp,
        width=24,
    )
    btn_mp.pack(side="left")

    tk.Entry(
        fila_mp,
        textvariable=archivo_mp_var,
        state="readonly",
    ).pack(side="left", fill="x", expand=True, padx=(10, 0))

    # Sucursales
    marco_suc = tk.LabelFrame(
        marco,
        text="Archivos de sucursales",
        padx=10,
        pady=10,
    )
    marco_suc.pack(fill="x", pady=10)

    fila_botones = tk.Frame(marco_suc)
    fila_botones.pack(fill="x")

    btn_suc = tk.Button(
        fila_botones,
        text="Agregar varias sucursales",
        command=seleccionar_sucursales,
        width=24,
    )
    btn_suc.pack(side="left")

    tk.Button(
        fila_botones,
        text="Agregar 1 sucursal",
        command=agregar_sucursal,
    ).pack(side="left", padx=8)

    tk.Button(
        fila_botones,
        text="Quitar seleccionada",
        command=quitar_sucursal,
    ).pack(side="left", padx=8)

    tk.Button(
        fila_botones,
        text="Limpiar",
        command=limpiar_sucursales,
    ).pack(side="left", padx=8)

    etiqueta_cantidad = tk.Label(
        fila_botones,
        text="0 archivo(s) de sucursal seleccionado(s)",
    )
    etiqueta_cantidad.pack(side="left")

    lista = tk.Listbox(marco_suc, height=5)
    lista.pack(fill="x", pady=(8, 0))

    # Botones inferiores
    # IMPORTANTE: este frame se empaqueta ANTES que la consola (más abajo).
    # Así, si la ventana queda con poco alto (pantallas chicas o con
    # escalado de Windows), Tkinter le garantiza espacio a los botones
    # primero y es la consola (expandible) la que se achica. Si se
    # empaqueta al revés, el botón puede quedar fuera del área visible.
    pie = tk.Frame(marco)
    pie.pack(fill="x", side="bottom", pady=(10, 0))

    btn_ejecutar = tk.Button(
        pie,
        text="▶  INICIAR CONCILIACIÓN",
        command=ejecutar_gui,
        font=("TkDefaultFont", 11, "bold"),
        padx=20,
        pady=8,
    )
    btn_ejecutar.pack(side="left")

    tk.Button(
        pie,
        text="Limpiar consola",
        command=limpiar_consola,
    ).pack(side="right")

    # Consola
    marco_consola = tk.LabelFrame(
        marco,
        text="Consola / resultado",
        padx=8,
        pady=8,
    )
    marco_consola.pack(fill="both", expand=True, pady=(5, 10))

    consola = scrolledtext.ScrolledText(
        marco_consola,
        state="disabled",
        wrap="none",
        font=("TkFixedFont", 9),
    )
    consola.pack(fill="both", expand=True)

    consola.tag_configure("verde", foreground="#008000")
    consola.tag_configure("rojo", foreground="#c00000")
    consola.tag_configure("amarillo", foreground="#b88600")
    consola.tag_configure("titulo", font=("TkFixedFont", 9, "bold"))

    root.mainloop()


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    # Si se ejecuta normalmente, abre la interfaz gráfica.
    iniciar_gui()